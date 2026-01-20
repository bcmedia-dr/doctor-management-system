from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
import os

def import_doctors_from_excel(file_path, db, Doctor):
    """從 Excel 文件匯入醫師資料（按照匯出格式）"""
    errors = []
    success_count = 0
    
    try:
        # 檢查檔案是否存在
        if not os.path.exists(file_path):
            return {
                'success': False,
                'success_count': 0,
                'errors': ['匯入失敗：檔案不存在']
            }
        
        # 檢查檔案大小（避免處理過大的檔案）
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            return {
                'success': False,
                'success_count': 0,
                'errors': ['匯入失敗：檔案為空']
            }
        if file_size > 10 * 1024 * 1024:  # 10MB
            return {
                'success': False,
                'success_count': 0,
                'errors': ['匯入失敗：檔案過大（超過 10MB）']
            }
        
        # 檢查檔案格式（通過副檔名和檔案內容）
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            return {
                'success': False,
                'success_count': 0,
                'errors': [f'匯入失敗：不支援的檔案格式 ({file_ext})。請使用 .xlsx、.xlsm、.xltx 或 .xltm 格式']
            }
        
        # 嘗試讀取檔案，檢查是否為有效的 Excel 檔案
        try:
            # 使用 data_only=True 來讀取計算後的值
            wb = load_workbook(file_path, data_only=True, read_only=True)
        except InvalidFileException as e:
            return {
                'success': False,
                'success_count': 0,
                'errors': [f'匯入失敗：檔案格式不正確。請確認檔案是有效的 Excel 檔案（.xlsx 格式），並且可以用 Excel 開啟。錯誤詳情：{str(e)}']
            }
        except Exception as e:
            error_msg = str(e)
            if 'does not support file format' in error_msg or 'file format' in error_msg.lower():
                return {
                    'success': False,
                    'success_count': 0,
                    'errors': ['匯入失敗：檔案格式不支援。請確認檔案是 .xlsx 格式，並且可以用 Excel 正常開啟。如果檔案是 .xls 格式，請先用 Excel 轉換為 .xlsx 格式。']
                }
            else:
                return {
                    'success': False,
                    'success_count': 0,
                    'errors': [f'匯入失敗：無法讀取檔案。請確認檔案沒有損壞，並且可以用 Excel 開啟。錯誤詳情：{str(e)}']
                }
        
        sheet = wb.active
        
        # 檢查工作表是否為空
        if sheet.max_row < 2:
            wb.close()
            return {
                'success': False,
                'success_count': 0,
                'errors': ['匯入失敗：Excel 檔案中沒有資料（至少需要標題列和一行資料）']
            }
        
        # 讀取標題行（第一行），確保正確處理中文編碼
        headers = []
        for cell in sheet[1]:
            if cell.value is not None:
                # 確保正確處理字符串編碼
                header_value = str(cell.value).strip()
                headers.append(header_value)
            else:
                headers.append('')
        
        # 定義匯出檔案的標準欄位順序（與 export.py 完全一致）
        # 匯出標題順序：醫師、科別、性別、狀態、聯絡窗口、合作品牌、報價區間、經營社群、醫師社群
        expected_headers = ['醫師', '科別', '性別', '狀態', '聯絡窗口', '合作品牌', '報價區間', '經營社群', '醫師社群']
        
        # 找到各欄位的索引位置（優先使用標準順序，如果標題匹配）
        column_map = {}
        use_standard_order = True
        
        # 檢查是否與標準格式一致
        if len(headers) >= len(expected_headers):
            matches = sum(1 for i, h in enumerate(expected_headers) if i < len(headers) and headers[i] == h)
            if matches >= 7:  # 至少7個欄位匹配，使用標準順序
                use_standard_order = True
                for idx, header_name in enumerate(expected_headers):
                    column_map[header_name] = idx
            else:
                use_standard_order = False
        
        # 如果不使用標準順序，則通過標題名稱查找（精確匹配優先）
        if not use_standard_order:
            for idx, header in enumerate(headers):
                header_str = str(header).strip() if header else ''
                # 精確匹配優先，避免誤判（按照標準欄位順序）
                if header_str == '醫師':
                    column_map['醫師'] = idx
                elif header_str == '科別':
                    column_map['科別'] = idx
                elif header_str == '性別':
                    column_map['性別'] = idx
                elif header_str == '狀態':
                    column_map['狀態'] = idx
                elif header_str == '聯絡窗口':
                    column_map['聯絡窗口'] = idx
                elif header_str == '經營社群':
                    column_map['經營社群'] = idx
                elif header_str == '醫師社群':
                    column_map['醫師社群'] = idx
                elif header_str == '合作品牌':
                    column_map['合作品牌'] = idx
                elif header_str == '報價區間':
                    column_map['報價區間'] = idx
                # 如果精確匹配失敗，使用模糊匹配作為備選（按照標準欄位順序）
                elif ('醫師' in header_str or 'Email' in header_str or 'email' in header_str.lower()) and '醫師' not in column_map:
                    column_map['醫師'] = idx
                elif '科別' in header_str and '科別' not in column_map:
                    column_map['科別'] = idx
                elif '性別' in header_str and '性別' not in column_map:
                    column_map['性別'] = idx
                elif '狀態' in header_str and '狀態' not in column_map:
                    column_map['狀態'] = idx
                elif '聯絡' in header_str and '窗口' in header_str and '聯絡窗口' not in column_map:
                    column_map['聯絡窗口'] = idx
                elif '經營社群' in header_str and '經營社群' not in column_map:
                    column_map['經營社群'] = idx
                elif '醫師社群' in header_str and '醫師社群' not in column_map:
                    column_map['醫師社群'] = idx
                elif '合作品牌' in header_str and '合作品牌' not in column_map:
                    column_map['合作品牌'] = idx
                elif '報價' in header_str and '報價區間' not in column_map:
                    column_map['報價區間'] = idx
        
        # 檢查必要的欄位是否存在（至少需要「醫師」欄位）
        required_fields = ['醫師']
        missing_fields = [field for field in required_fields if field not in column_map]
        if missing_fields:
            wb.close()
            return {
                'success': False,
                'success_count': 0,
                'errors': [f'匯入失敗：Excel 檔案中缺少必要的欄位：{", ".join(missing_fields)}。請確認檔案格式正確。']
            }
        
        # 從第二行開始讀取資料
        batch_size = 50  # 每批處理50筆資料
        doctors_to_add = []
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            # 跳過空行
            if not any(cell.value for cell in row):
                continue
            
            try:
                # 提取資料的輔助函數
                def get_cell_value(header_name, default_idx=None):
                    idx = column_map.get(header_name, default_idx)
                    if idx is None:
                        return ''
                    if idx < len(row) and row[idx].value is not None:
                        value = row[idx].value
                        # 處理不同類型的值
                        if isinstance(value, (int, float)):
                            # 數字類型直接轉字符串
                            if isinstance(value, float) and value.is_integer():
                                return str(int(value))
                            return str(value)
                        elif isinstance(value, datetime):
                            # 日期時間類型轉字符串
                            return value.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            # 字符串類型，確保正確編碼
                            return str(value).strip()
                    return ''
                
                # 按照標準欄位順序讀取：醫師、科別、性別、狀態、聯絡窗口、經營社群、醫師社群、合作品牌、報價區間
                doctor_field = get_cell_value('醫師', 0).strip() if get_cell_value('醫師', 0) else ''
                
                # 如果醫師欄位為空，跳過這一行
                if not doctor_field:
                    continue
                
                # 同時存到 name 和 email
                name = doctor_field
                email = doctor_field
                
                # 清理和驗證資料（按照標準欄位順序：醫師、科別、性別、狀態、聯絡窗口、合作品牌、報價區間、經營社群、醫師社群）
                specialty = get_cell_value('科別', 1).strip() if get_cell_value('科別', 1) else None
                gender = get_cell_value('性別', 2).strip() if get_cell_value('性別', 2) else None
                status = (get_cell_value('狀態', 3) or '未聯繫').strip()
                contact_person = get_cell_value('聯絡窗口', 4).strip() if get_cell_value('聯絡窗口', 4) else None
                current_brand = get_cell_value('合作品牌', 5).strip() if get_cell_value('合作品牌', 5) else None
                price_range = get_cell_value('報價區間', 6).strip() if get_cell_value('報價區間', 6) else None
                has_social_media = get_cell_value('經營社群', 7).strip() if get_cell_value('經營社群', 7) else None
                social_media_link = get_cell_value('醫師社群', 8).strip() if get_cell_value('醫師社群', 8) else None
                
                # 檢查是否已存在相同email或name的醫師（避免重複匯入）
                existing_doctor = Doctor.query.filter(
                    (Doctor.email == email) | (Doctor.name == name)
                ).first()
                if existing_doctor:
                    errors.append(f"第 {row_num} 行：醫師 '{name or email}' 已存在，已跳過")
                    continue
                
                # 創建新醫師記錄（按照標準欄位順序）
                doctor = Doctor(
                    name=name if name else None,  # 姓名
                    email=email if email else None,  # Email
                    specialty=specialty if specialty else None,  # 科別
                    gender=gender if gender else None,  # 性別
                    status=status if status else '未聯繫',  # 狀態
                    contact_person=contact_person if contact_person else None,  # 聯絡窗口
                    has_social_media=has_social_media if has_social_media else None,  # 經營社群
                    social_media_link=social_media_link if social_media_link else None,  # 醫師社群
                    current_brand=current_brand if current_brand else None,  # 合作品牌
                    price_range=price_range if price_range else None  # 報價區間
                )
                
                doctors_to_add.append(doctor)
                success_count += 1
                
                # 批量提交以提高效率
                if len(doctors_to_add) >= batch_size:
                    try:
                        for doc in doctors_to_add:
                            db.session.add(doc)
                        db.session.commit()
                        doctors_to_add = []
                    except Exception as batch_error:
                        db.session.rollback()
                        errors.append(f"第 {row_num} 行批量提交失敗：{str(batch_error)}")
                        # 嘗試逐筆添加
                        for doc in doctors_to_add:
                            try:
                                db.session.add(doc)
                                db.session.commit()
                            except Exception as single_error:
                                db.session.rollback()
                                errors.append(f"醫師 '{doc.email}' 添加失敗：{str(single_error)}")
                                success_count -= 1  # 扣除失敗的計數
                        doctors_to_add = []
                
            except Exception as e:
                errors.append(f"第 {row_num} 行處理失敗：{str(e)}")
                continue
        
        # 提交剩餘的資料
        if doctors_to_add:
            try:
                for doc in doctors_to_add:
                    db.session.add(doc)
                db.session.commit()
            except Exception as batch_error:
                db.session.rollback()
                errors.append(f"批量提交剩餘資料失敗：{str(batch_error)}")
                # 嘗試逐筆添加
                for doc in doctors_to_add:
                    try:
                        db.session.add(doc)
                        db.session.commit()
                    except Exception as single_error:
                        db.session.rollback()
                        errors.append(f"醫師 '{doc.email}' 添加失敗：{str(single_error)}")
                        success_count -= 1  # 扣除失敗的計數
        
        # 關閉工作簿
        wb.close()
        
        return {
            'success': True,
            'success_count': success_count,
            'errors': errors
        }
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        
        # 提供更友好的錯誤訊息
        if 'does not support file format' in error_msg or 'file format' in error_msg.lower():
            friendly_error = '匯入失敗：檔案格式不支援。請確認檔案是 .xlsx 格式，並且可以用 Excel 正常開啟。如果檔案是 .xls 格式，請先用 Excel 轉換為 .xlsx 格式。'
        elif 'No such file' in error_msg or 'cannot find' in error_msg.lower():
            friendly_error = '匯入失敗：找不到檔案。請確認檔案已正確上傳。'
        elif 'Permission denied' in error_msg:
            friendly_error = '匯入失敗：沒有權限讀取檔案。'
        else:
            friendly_error = f'匯入失敗：{error_msg}'
        
        return {
            'success': False,
            'success_count': 0,
            'errors': [friendly_error]
        }
