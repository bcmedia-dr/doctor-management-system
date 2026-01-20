from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def export_doctors_to_excel(doctors):
    """匯出醫師資料到 Excel"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "醫師資料"
    
    # 設定標題列（簡化欄位，只保留必要資訊）
    # 匯出順序：醫師、科別、性別、狀態、聯絡窗口、經營社群、醫師社群、合作品牌、報價區間
    headers = ['醫師', '科別', '性別', '狀態', '聯絡窗口', '經營社群', '醫師社群', '合作品牌', '報價區間']
    sheet.append(headers)
    
    # 設定標題列樣式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 寫入資料（優先使用 email，沒有就用 name）
    for index, doctor in enumerate(doctors, 1):
        doctor_name = doctor.name or ''
        doctor_email = doctor.email or ''
        # 醫師欄位：優先顯示 email，沒有就用 name
        doctor_field = doctor_email or doctor_name or ''
        row_data = [
            doctor_field,  # 醫師：優先使用 email，沒有就用 name
            doctor.specialty or '',  # 科別
            doctor.gender or '',  # 性別
            doctor.status or '',  # 狀態
            doctor.contact_person or '',  # 聯絡窗口
            doctor.has_social_media or '',  # 經營社群
            doctor.social_media_link or '',  # 醫師社群
            doctor.current_brand or '',  # 合作品牌
            doctor.price_range or ''  # 報價區間
        ]
        sheet.append(row_data)
    
    # 調整欄寬（按照新的欄位順序）
    column_widths = {
        'A': 30,  # 醫師
        'B': 12,  # 科別
        'C': 8,   # 性別
        'D': 12,  # 狀態
        'E': 12,  # 聯絡窗口
        'F': 12,  # 經營社群
        'G': 30,  # 醫師社群
        'H': 20,  # 合作品牌
        'I': 15   # 報價區間
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # 設定資料列樣式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:  # 資料列
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 凍結首列
    sheet.freeze_panes = 'A2'
    
    # 儲存檔案（確保UTF-8編碼支持）
    file_path = f'/tmp/doctors_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    # openpyxl 自動處理 UTF-8 編碼，無需額外設置
    wb.save(file_path)
    
    return file_path
