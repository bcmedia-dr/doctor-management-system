from openpyxl import load_workbook
from datetime import datetime

def import_doctors_from_excel(file_path, db, Doctor):
    """從 Excel 文件匯入醫師資料"""
    errors = []
    success_count = 0
    
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
        
        # 讀取標題行（第一行）
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value if cell.value else '')
        
        # 找到各欄位的索引位置
        column_map = {}
        for idx, header in enumerate(headers):
            header_str = str(header).strip() if header else ''
            if '編號' in header_str or '序號' in header_str:
                column_map['index'] = idx
            elif '醫師' in header_str:
                column_map['email'] = idx
            elif '科別' in header_str:
                column_map['specialty'] = idx
            elif '性別' in header_str:
                column_map['gender'] = idx
            elif '狀態' in header_str:
                column_map['status'] = idx
            elif '聯絡' in header_str and '窗口' in header_str:
                column_map['contact_person'] = idx
            elif '合作品牌' in header_str:
                column_map['current_brand'] = idx
            elif '報價' in header_str:
                column_map['price_range'] = idx
            elif '經營社群' in header_str:
                column_map['has_social_media'] = idx
            elif '醫師社群' in header_str or '社群' in header_str:
                column_map['social_media_link'] = idx
        
        # 從第二行開始讀取資料
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            # 跳過空行
            if not any(cell.value for cell in row):
                continue
            
            try:
                # 提取資料
                email = str(row[column_map.get('email', 1)].value).strip() if column_map.get('email') and row[column_map['email']].value else ''
                
                # 如果醫師欄位為空，跳過這一行
                if not email:
                    continue
                
                specialty = str(row[column_map.get('specialty', 2)].value).strip() if column_map.get('specialty') and row[column_map['specialty']].value else ''
                gender = str(row[column_map.get('gender', 3)].value).strip() if column_map.get('gender') and row[column_map['gender']].value else ''
                status = str(row[column_map.get('status', 4)].value).strip() if column_map.get('status') and row[column_map['status']].value else '未聯繫'
                contact_person = str(row[column_map.get('contact_person', 5)].value).strip() if column_map.get('contact_person') and row[column_map['contact_person']].value else ''
                current_brand = str(row[column_map.get('current_brand', 6)].value).strip() if column_map.get('current_brand') and row[column_map['current_brand']].value else ''
                price_range = str(row[column_map.get('price_range', 7)].value).strip() if column_map.get('price_range') and row[column_map['price_range']].value else ''
                has_social_media = str(row[column_map.get('has_social_media', 8)].value).strip() if column_map.get('has_social_media') and row[column_map['has_social_media']].value else ''
                social_media_link = str(row[column_map.get('social_media_link', 9)].value).strip() if column_map.get('social_media_link') and row[column_map['social_media_link']].value else ''
                
                # 檢查是否已存在相同email的醫師（避免重複匯入）
                existing_doctor = Doctor.query.filter_by(email=email).first()
                if existing_doctor:
                    errors.append(f"第 {row_num} 行：醫師 '{email}' 已存在，已跳過")
                    continue
                
                # 創建新醫師記錄
                doctor = Doctor(
                    name=email,  # 使用email字段存储醫師名称
                    email=email,
                    specialty=specialty if specialty else None,
                    gender=gender if gender else None,
                    status=status if status else '未聯繫',
                    contact_person=contact_person if contact_person else None,
                    has_social_media=has_social_media if has_social_media else None,
                    social_media_link=social_media_link if social_media_link else None,
                    current_brand=current_brand if current_brand else None,
                    price_range=price_range if price_range else None
                )
                
                db.session.add(doctor)
                success_count += 1
                
            except Exception as e:
                errors.append(f"第 {row_num} 行處理失敗：{str(e)}")
                continue
        
        # 提交所有變更
        db.session.commit()
        
        return {
            'success': True,
            'success_count': success_count,
            'errors': errors
        }
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': False,
            'success_count': 0,
            'errors': [f'匯入失敗：{str(e)}']
        }
